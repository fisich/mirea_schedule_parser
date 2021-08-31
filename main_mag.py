import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import os

# TODO : Рендер в изображение

#print('Название файла')
# filename = input()
filename = "ИИТ_маг_1к_21-22_осень.xlsx"
#print('Введите номер группы')
groupname = input()
if not os.path.isfile(filename):
    print("Неверное имя файла или файл не существует")
    exit()

workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

rownum = None
colnum = None
for row in sheet.iter_rows(2):
    for cell in row:
        value = str(cell.value)
        if value.find(groupname) > -1:
            rownum = cell.row
            colnum = cell.column
            break

if rownum is None:
    print('Группа не найдена')
    exit()

finalWorkBook = Workbook()
oddSheet = finalWorkBook.active
oddSheet.title = "Нечетная неделя"
#Заголовки
oddSheet.merge_cells('A1:A2')
oddSheet.merge_cells('B1:J1')

column_widths = [49, 39, 55, 53, 64, 101, 56, 64, 64, 64]

for i, column_width in enumerate(column_widths):
    oddSheet.column_dimensions[get_column_letter(i + 1)].width = column_width / 7

#56 - количество строк на выходе 6 дней * количество предметов + 2 строки из заголовков
for i in range(2, 57):
    oddSheet.merge_cells('E' + str(i) + ':F' + str(i))
    oddSheet.merge_cells('H' + str(i) + ':I' + str(i))
    if i > 2:
        oddSheet['B' + str(i)] = 1 + (i - 3) % 9 # проставляем номер занятия

for row in oddSheet.iter_rows(1, 56, 1, 10):
    for cell in row:
        #Края ячеек
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        cell.font = Font(name='TimesNewRoman')

oddSheet['A1'] = 'День недели'
oddSheet['B2'] = '№ пары'
oddSheet['C2'] = 'Нач. занятий'
oddSheet['D2'] = 'Оконч. занятий'
oddSheet['E2'] = 'Предмет'
oddSheet['G2'] = 'Вид занятий'
oddSheet['H2'] = 'ФИО преподавателя'
oddSheet['J2'] = '№ ауд.'

timelist = ()

# Инфа по расписанию
for row in sheet.iter_rows(4, 20, 3, 4):
    for cell in row:
        if cell.value:
            timelist = timelist + (str(cell.value),)

# Записываем инфу по расписанию
for i in range(0, 6):
    for j in range(0, 9):
        oddSheet['C' + str(j + 3 + i * 9)] = timelist[j * 2]
        oddSheet['D' + str(j + 3 + i * 9)] = timelist[j * 2 + 1]

# Склеиваем названия дней
oddSheet.merge_cells('A3:A11')
oddSheet.merge_cells('A12:A20')
oddSheet.merge_cells('A21:A29')
oddSheet.merge_cells('A30:A38')
oddSheet.merge_cells('A39:A47')
oddSheet.merge_cells('A48:A56')

days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
colors = ['E6B8B7', 'FFC000', '92D050', 'E26B10', '00B0F0', 'B1A0C7']

for i, day in enumerate(days):
    oddSheet['A' + str(3 + i * 9)] = day
    oddSheet['A' + str(3 + i * 9)].alignment = Alignment(text_rotation=90, vertical='center')
    oddSheet['A' + str(3 + i * 9)].fill = PatternFill(start_color=colors[i], fill_type="solid")
    oddSheet['A' + str(3 + i * 9)].font = Font(bold=True, size=18 if i else 13, name='TimesNewRoman')

evenSheet = finalWorkBook.copy_worksheet(oddSheet)
evenSheet.title = 'Четная неделя'

index = 3
isOdd = True

# Диапазон, пробегая через который заполняем расписание 2 - отступ заголовка из исходника
# 103 - последние строки в расписании (суббота - 9 пара для магов)
for row in sheet.iter_rows(min_row=rownum + 2, min_col=colnum, max_row=rownum + 103, max_col=colnum + 3):
    if isOdd:
        oddSheet['E' + str(index)] = row[0].value
        # Ручной подгон высоты ячейки
        if len(str(row[0].value)) > 10:
            oddSheet.row_dimensions[oddSheet['E' + str(index)].row].height = 30
        if len(str(row[0].value)) > 28:
            oddSheet['E' + str(index)].font = Font(name='TimesNewRoman', size=9)
        if len(str(row[0].value)) > 55:
            oddSheet.row_dimensions[oddSheet['E' + str(index)].row].height = 40
        if len(str(row[0].value)) > 90:
            oddSheet.row_dimensions[oddSheet['E' + str(index)].row].height = 50
        oddSheet['G' + str(index)] = row[1].value
        oddSheet['H' + str(index)] = row[2].value
        oddSheet['J' + str(index)] = row[3].value
        isOdd = False
    else:
        evenSheet['E' + str(index)] = row[0].value
        if len(str(row[0].value)) > 10:
            evenSheet.row_dimensions[evenSheet['E' + str(index)].row].height = 30
        if len(str(row[0].value)) > 28:
            evenSheet['E' + str(index)].font = Font(name='TimesNewRoman', size=9)
        if len(str(row[0].value)) > 55:
            evenSheet.row_dimensions[evenSheet['E' + str(index)].row].height = 40
        if len(str(row[0].value)) > 90:
            evenSheet.row_dimensions[evenSheet['E' + str(index)].row].height = 50
        evenSheet['G' + str(index)] = row[1].value
        evenSheet['H' + str(index)] = row[2].value
        evenSheet['J' + str(index)] = row[3].value
        index += 1
        isOdd = True

oddSheet['B1'] = groupname + '. ' + oddSheet.title
evenSheet['B1'] = groupname + '. ' + evenSheet.title

# Закрашиваем синим
fillBlue = [oddSheet['A1:J2'], oddSheet['B12:J20'], oddSheet['B30:J38'], oddSheet['B48:J56']]
for group in fillBlue:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="95B3D7", fill_type="solid")

# Закрашиваем голубым
fillBlue = [oddSheet['B3:J11'], oddSheet['B21:J29'], oddSheet['B39:J47']]
for group in fillBlue:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="B8CCE4", fill_type="solid")

# Закрашиваем красным
fillRed = [evenSheet['A1:J2'], evenSheet['B12:J20'], evenSheet['B30:J38'], evenSheet['B48:J56']]
for group in fillRed:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="FABF8F", fill_type="solid")

# Закрашиваем оранжевым
fillRed = [evenSheet['B3:J11'], evenSheet['B21:J29'], evenSheet['B39:J47']]
for group in fillRed:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="FDE9D9", fill_type="solid")


finalWorkBook.save('Расписание ' + groupname + '.xlsx')
