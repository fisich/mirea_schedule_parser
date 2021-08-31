import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import os

#print('Название файла')
# filename = input()
filename = "ИИТ_маг_1к_21-22_осень.xlsx"
print('Введите номер группы')
groupname = input()
if not os.path.isfile(filename):
    print("Неверное имя файла или файл не существует")
    exit()

workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

rownum = None
colnum = None
for row in sheet.iter_rows(2):
    for cell in row:
        value = str(cell.value)
        if value.find(groupname) > -1:
            rownum = cell.row
            colnum = cell.column
            break

if rownum is None:
    print('Группа не найдена')
    exit()

finalWorkBook = Workbook()
oddSheet = finalWorkBook.active
oddSheet.title = "Нечетная неделя"
oddSheet.merge_cells('A1:A2')
oddSheet.merge_cells('B1:J1')

column_widths = [49, 39, 55, 53, 64, 101, 56, 64, 64, 64]

for i, column_width in enumerate(column_widths):
    oddSheet.column_dimensions[get_column_letter(i + 1)].width = column_width / 7

for i in range(2, 39):
    oddSheet.merge_cells('E' + str(i) + ':F' + str(i))
    oddSheet.merge_cells('H' + str(i) + ':I' + str(i))
    if i > 2:
        oddSheet['B' + str(i)] = 1 + (i - 3) % 6

for row in oddSheet.iter_rows(1, 38, 1, 10):
    for cell in row:
        cell.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        cell.font = Font(name='TimesNewRoman')

oddSheet['A1'] = 'День недели'
oddSheet['B2'] = '№ пары'
oddSheet['C2'] = 'Нач. занятий'
oddSheet['D2'] = 'Оконч. занятий'
oddSheet['E2'] = 'Предмет'
oddSheet['G2'] = 'Вид занятий'
oddSheet['H2'] = 'ФИО преподавателя'
oddSheet['J2'] = '№ ауд.'

timelist = ()

for row in sheet.iter_rows(4, 14, 3, 4):
    for cell in row:
        if cell.value:
            timelist = timelist + (str(cell.value),)

for i in range(0, 6):
    for j in range(0, 6):
        oddSheet['C' + str(j + 3 + i * 6)] = timelist[j * 2]
        oddSheet['D' + str(j + 3 + i * 6)] = timelist[j * 2 + 1]

oddSheet.merge_cells('A3:A8')
oddSheet.merge_cells('A9:A14')
oddSheet.merge_cells('A15:A20')
oddSheet.merge_cells('A21:A26')
oddSheet.merge_cells('A27:A32')
oddSheet.merge_cells('A33:A38')

days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
colors = ['E6B8B7', 'FFC000', '92D050', 'E26B10', '00B0F0', 'B1A0C7']

for i, day in enumerate(days):
    oddSheet['A' + str(3 + i * 6)] = day
    oddSheet['A' + str(3 + i * 6)].alignment = Alignment(text_rotation=90, vertical='center')
    oddSheet['A' + str(3 + i * 6)].fill = PatternFill(start_color=colors[i], fill_type="solid")
    oddSheet['A' + str(3 + i * 6)].font = Font(bold=True, size=18 if i else 13, name='TimesNewRoman')

evenSheet = finalWorkBook.copy_worksheet(oddSheet)
evenSheet.title = 'Четная неделя'

index = 3
isOdd = True

for row in sheet.iter_rows(min_row=rownum + 2, min_col=colnum, max_row=rownum + 73, max_col=colnum + 3):
    if isOdd:
        oddSheet['E' + str(index)] = row[0].value
        if len(str(row[0].value)) > 20:
            oddSheet['E' + str(index)].font = Font(name='TimesNewRoman', size=9)
            oddSheet.row_dimensions[oddSheet['E' + str(index)].row].height = 30
        oddSheet['G' + str(index)] = row[1].value
        oddSheet['H' + str(index)] = row[2].value
        oddSheet['J' + str(index)] = row[3].value
        isOdd = False
    else:
        evenSheet['E' + str(index)] = row[0].value
        if len(str(row[0].value)) > 20:
            evenSheet['E' + str(index)].font = Font(name='TimesNewRoman', size=9)
            evenSheet.row_dimensions[evenSheet['E' + str(index)].row].height = 30
        evenSheet['G' + str(index)] = row[1].value
        evenSheet['H' + str(index)] = row[2].value
        evenSheet['J' + str(index)] = row[3].value
        index += 1
        isOdd = True

oddSheet['B1'] = groupname + '. ' + oddSheet.title
evenSheet['B1'] = groupname + '. ' + evenSheet.title

fillBlue = [oddSheet['A1:J2'], oddSheet['B9:J14'], oddSheet['B21:J26'], oddSheet['B33:J38']]
for group in fillBlue:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="95B3D7", fill_type="solid")

fillBlue = [oddSheet['B3:J8'], oddSheet['B15:J20'], oddSheet['B27:J32']]
for group in fillBlue:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="B8CCE4", fill_type="solid")

fillRed = [evenSheet['A1:J2'], evenSheet['B9:J14'], evenSheet['B21:J26'], evenSheet['B33:J38']]
for group in fillRed:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="FABF8F", fill_type="solid")

fillRed = [evenSheet['B3:J8'], evenSheet['B15:J20'], evenSheet['B27:J32']]
for group in fillRed:
    for row in group:
        for cell in row:
            cell.fill = PatternFill(start_color="FDE9D9", fill_type="solid")


finalWorkBook.save('Расписание ' + groupname + '.xlsx')
